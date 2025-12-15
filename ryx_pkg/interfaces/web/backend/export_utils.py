"""
Export Utilities - Generate PDF/Excel/JSON reports

Supports multiple export formats for grading results with full pedagogical
feedback, learning analytics, and visualizations.
"""

import logging
import json
from typing import Dict, Any, Optional, List
from datetime import datetime
from io import BytesIO

logger = logging.getLogger(__name__)

# Optional dependencies
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab not installed - PDF export disabled")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel export disabled")


def export_grading_result(
    grading_result: Dict[str, Any],
    mock_exam: Dict[str, Any],
    format: str = "json"
) -> bytes:
    """
    Export grading result in specified format.
    
    Args:
        grading_result: Full grading result dict
        mock_exam: Exam metadata
        format: 'json', 'pdf', or 'excel'
    
    Returns:
        Bytes of exported file
    """
    
    if format == "json":
        return _export_json(grading_result, mock_exam)
    elif format == "pdf" and PDF_AVAILABLE:
        return _export_pdf(grading_result, mock_exam)
    elif format == "excel" and EXCEL_AVAILABLE:
        return _export_excel(grading_result, mock_exam)
    else:
        raise ValueError(f"Format '{format}' not available or not supported")


def _export_json(grading_result: Dict[str, Any], mock_exam: Dict[str, Any]) -> bytes:
    """Export as structured JSON"""
    
    export_data = {
        "exam_metadata": {
            "exam_id": mock_exam.get("id"),
            "title": mock_exam.get("title", "Unbenannt"),
            "subject": mock_exam.get("subject_name", "N/A"),
            "thema": mock_exam.get("main_thema", "N/A"),
            "difficulty": mock_exam.get("difficulty", 3),
            "total_tasks": len(mock_exam.get("tasks", [])),
        },
        "grading_result": grading_result,
        "exported_at": datetime.utcnow().isoformat(),
        "export_format": "json",
        "version": "2.0"
    }
    
    return json.dumps(export_data, indent=2, ensure_ascii=False).encode("utf-8")


def _export_pdf(grading_result: Dict[str, Any], mock_exam: Dict[str, Any]) -> bytes:
    """Export as formatted PDF report"""
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor("#2c3e50"), alignment=TA_CENTER)
    heading_style = styles['Heading2']
    normal_style = styles['Normal']
    
    story = []
    
    # Title
    story.append(Paragraph("Bewertungsbericht", title_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Exam metadata
    story.append(Paragraph(f"<b>Prüfung:</b> {mock_exam.get('title', 'Unbenannt')}", normal_style))
    story.append(Paragraph(f"<b>Fach:</b> {mock_exam.get('subject_name', 'N/A')}", normal_style))
    story.append(Paragraph(f"<b>Thema:</b> {mock_exam.get('main_thema', 'N/A')}", normal_style))
    story.append(Paragraph(f"<b>Datum:</b> {datetime.fromisoformat(grading_result['created_at']).strftime('%d.%m.%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Overall result
    story.append(Paragraph("Gesamtergebnis", heading_style))
    result_data = [
        ["Erreichte Punkte", f"{grading_result['total_score']:.1f} / {grading_result['total_points']}"],
        ["Prozent", f"{grading_result['percentage']:.1f}%"],
        ["Note", f"{grading_result['grade']} ({grading_result['grade_text']})"],
        ["KI-Konfidenz", f"{grading_result['grader_confidence']}%"],
    ]
    
    result_table = Table(result_data, colWidths=[8*cm, 8*cm])
    result_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#ecf0f1")),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    story.append(result_table)
    story.append(Spacer(1, 0.5*cm))
    
    # Overall feedback
    story.append(Paragraph("Gesamtfeedback", heading_style))
    feedback_text = grading_result.get('overall_feedback', 'Kein Feedback verfügbar.')
    story.append(Paragraph(feedback_text.replace('\n', '<br/>'), normal_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Learning analytics (if available)
    learning_analytics = grading_result.get('learning_analytics')
    if learning_analytics:
        story.append(PageBreak())
        story.append(Paragraph("Lernanalyse", heading_style))
        
        # Strengths
        strengths = learning_analytics.get('strengths', [])
        if strengths:
            story.append(Paragraph("<b>Stärken:</b>", normal_style))
            for strength in strengths:
                story.append(Paragraph(f"• {strength}", normal_style))
            story.append(Spacer(1, 0.3*cm))
        
        # Weaknesses
        weaknesses = learning_analytics.get('weaknesses', [])
        if weaknesses:
            story.append(Paragraph("<b>Verbesserungsbereiche:</b>", normal_style))
            for weakness in weaknesses:
                story.append(Paragraph(f"• {weakness}", normal_style))
            story.append(Spacer(1, 0.3*cm))
        
        # Topic masteries
        topic_masteries = learning_analytics.get('topic_masteries', [])
        if topic_masteries:
            story.append(Paragraph("Themenbeherrschung", heading_style))
            topic_data = [["Thema", "Beherrschung", "Level"]]
            for tm in topic_masteries[:10]:  # Top 10
                topic_data.append([
                    tm['topic_name'][:30],
                    f"{tm['mastery_percentage']:.0f}%",
                    tm['mastery_level']
                ])
            
            topic_table = Table(topic_data, colWidths=[10*cm, 3*cm, 3*cm])
            topic_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#3498db")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            story.append(topic_table)
            story.append(Spacer(1, 0.5*cm))
        
        # Recommendations
        recommendations = learning_analytics.get('recommendations', [])
        if recommendations:
            story.append(Paragraph("Lernempfehlungen", heading_style))
            for rec in recommendations[:5]:  # Top 5
                story.append(Paragraph(f"<b>{rec['priority']}. {rec['topic']}</b>", normal_style))
                story.append(Paragraph(f"Grund: {rec['reason']}", normal_style))
                story.append(Paragraph(f"Maßnahme: {rec['action']}", normal_style))
                if rec.get('resource'):
                    story.append(Paragraph(f"Ressource: {rec['resource']}", normal_style))
                story.append(Spacer(1, 0.3*cm))
    
    # Task breakdown
    story.append(PageBreak())
    story.append(Paragraph("Aufgabenbewertung", heading_style))
    
    for tg in grading_result.get('task_grades', []):
        story.append(Paragraph(f"<b>Aufgabe {tg['task_id']}</b>", normal_style))
        story.append(Paragraph(f"Punkte: {tg['earned_points']:.1f} / {tg['max_points']}", normal_style))
        story.append(Paragraph(f"Begründung: {tg['rationale']}", normal_style))
        
        # Pedagogical feedback
        ped_feedback = tg.get('pedagogical_feedback')
        if ped_feedback:
            if ped_feedback.get('what_was_good'):
                story.append(Paragraph(f"✓ Gut: {ped_feedback['what_was_good']}", normal_style))
            if ped_feedback.get('what_was_missing'):
                story.append(Paragraph(f"⚠ Fehlt: {ped_feedback['what_was_missing']}", normal_style))
        
        story.append(Spacer(1, 0.3*cm))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def _export_excel(grading_result: Dict[str, Any], mock_exam: Dict[str, Any]) -> bytes:
    """Export as Excel spreadsheet with multiple sheets"""
    
    buffer = BytesIO()
    wb = openpyxl.Workbook()
    
    # Sheet 1: Overview
    ws_overview = wb.active
    ws_overview.title = "Übersicht"
    
    # Header styling
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    ws_overview['A1'] = "Bewertungsbericht"
    ws_overview['A1'].font = Font(bold=True, size=16)
    ws_overview.merge_cells('A1:B1')
    
    ws_overview['A3'] = "Prüfung"
    ws_overview['B3'] = mock_exam.get('title', 'Unbenannt')
    ws_overview['A4'] = "Fach"
    ws_overview['B4'] = mock_exam.get('subject_name', 'N/A')
    ws_overview['A5'] = "Thema"
    ws_overview['B5'] = mock_exam.get('main_thema', 'N/A')
    ws_overview['A6'] = "Datum"
    ws_overview['B6'] = datetime.fromisoformat(grading_result['created_at']).strftime('%d.%m.%Y %H:%M')
    
    ws_overview['A8'] = "Erreichte Punkte"
    ws_overview['B8'] = f"{grading_result['total_score']:.1f} / {grading_result['total_points']}"
    ws_overview['A9'] = "Prozent"
    ws_overview['B9'] = f"{grading_result['percentage']:.1f}%"
    ws_overview['A10'] = "Note"
    ws_overview['B10'] = f"{grading_result['grade']} ({grading_result['grade_text']})"
    
    # Bold first column
    for row in range(3, 11):
        ws_overview[f'A{row}'].font = Font(bold=True)
    
    # Sheet 2: Task breakdown
    ws_tasks = wb.create_sheet("Aufgaben")
    ws_tasks['A1'] = "Aufgaben-ID"
    ws_tasks['B1'] = "Typ"
    ws_tasks['C1'] = "Erreicht"
    ws_tasks['D1'] = "Maximum"
    ws_tasks['E1'] = "Prozent"
    ws_tasks['F1'] = "Begründung"
    
    # Style header
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_tasks[f'{col}1'].fill = header_fill
        ws_tasks[f'{col}1'].font = header_font
    
    # Fill task data
    row = 2
    for tg in grading_result.get('task_grades', []):
        ws_tasks[f'A{row}'] = tg['task_id']
        ws_tasks[f'B{row}'] = tg['task_type']
        ws_tasks[f'C{row}'] = tg['earned_points']
        ws_tasks[f'D{row}'] = tg['max_points']
        ws_tasks[f'E{row}'] = f"{(tg['earned_points'] / tg['max_points'] * 100):.1f}%"
        ws_tasks[f'F{row}'] = tg['rationale'][:100]  # Truncate long text
        row += 1
    
    # Adjust column widths
    ws_tasks.column_dimensions['A'].width = 15
    ws_tasks.column_dimensions['B'].width = 20
    ws_tasks.column_dimensions['C'].width = 12
    ws_tasks.column_dimensions['D'].width = 12
    ws_tasks.column_dimensions['E'].width = 12
    ws_tasks.column_dimensions['F'].width = 50
    
    # Sheet 3: Learning Analytics (if available)
    learning_analytics = grading_result.get('learning_analytics')
    if learning_analytics:
        ws_analytics = wb.create_sheet("Lernanalyse")
        ws_analytics['A1'] = "Thema"
        ws_analytics['B1'] = "Beherrschung %"
        ws_analytics['C1'] = "Level"
        ws_analytics['D1'] = "Fragen"
        
        for col in ['A', 'B', 'C', 'D']:
            ws_analytics[f'{col}1'].fill = header_fill
            ws_analytics[f'{col}1'].font = header_font
        
        row = 2
        for tm in learning_analytics.get('topic_masteries', []):
            ws_analytics[f'A{row}'] = tm['topic_name']
            ws_analytics[f'B{row}'] = tm['mastery_percentage']
            ws_analytics[f'C{row}'] = tm['mastery_level']
            ws_analytics[f'D{row}'] = tm['questions_answered']
            row += 1
        
        ws_analytics.column_dimensions['A'].width = 30
        ws_analytics.column_dimensions['B'].width = 15
        ws_analytics.column_dimensions['C'].width = 15
        ws_analytics.column_dimensions['D'].width = 12
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
